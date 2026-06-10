from __future__ import annotations

from copy import copy
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


OPEN_RECORD_HEADERS = [
    "序号",
    "实验室名称及位置",
    "开放实验（训）项目名称",
    "类型",
    "实验（训）人数",
    "开放实验（训）日期+时间",
    "学时",
    "实验（训）仪器设备使用情况",
    "项目负责人签字",
    "使用单位",
]


def infer_project_type(purpose: str) -> str:
    text = purpose or ""
    if "竞赛" in text or "集训" in text:
        return "竞赛"
    if "科研" in text or "课题" in text or "研究" in text:
        return "科研"
    if "课程" in text or "教学" in text or "实验" in text:
        return "课程/教学"
    return "开放预约"


def duration_credit_hours(start_dt: datetime | None, end_dt: datetime | None) -> float:
    if not start_dt or not end_dt or end_dt <= start_dt:
        return 0.0
    return round((end_dt - start_dt).total_seconds() / 60 / 45, 1)


def format_datetime_range(start_dt: datetime | None, end_dt: datetime | None) -> str:
    if not start_dt or not end_dt:
        return ""
    if start_dt.date() == end_dt.date():
        return f"{start_dt:%Y-%m-%d %H:%M}-{end_dt:%H:%M}"
    return f"{start_dt:%Y-%m-%d %H:%M} 至 {end_dt:%Y-%m-%d %H:%M}"


def reservation_to_open_record(index: int, item: dict) -> list[Any]:
    lab_display = item.get("lab_name", "")
    if item.get("lab_location"):
        lab_display = f"{lab_display}（{item.get('lab_location')}）"
    purpose = item.get("purpose", "")
    return [
        index,
        lab_display,
        purpose,
        infer_project_type(purpose),
        item.get("participant_count") or 1,
        format_datetime_range(item.get("start_dt"), item.get("end_dt")),
        duration_credit_hours(item.get("start_dt"), item.get("end_dt")),
        item.get("device_name") or "无",
        item.get("requester_name", item.get("requester_username", "")),
        item.get("requester_department", ""),
    ]


def _find_open_record_sheet(wb):
    for ws in wb.worksheets:
        if str(ws.cell(row=1, column=1).value or "").strip() == "实验（训）室开放记录":
            return ws
    return wb.active


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def export_open_records_with_template(template_bytes: bytes, reservations: list[dict]) -> bytes:
    wb = load_workbook(BytesIO(template_bytes))
    ws = _find_open_record_sheet(wb)
    data_start = 4
    max_col = 10
    style_row = 4 if ws.max_row >= 4 else 3

    if ws.max_row >= data_start:
        ws.delete_rows(data_start, ws.max_row - data_start + 1)

    if reservations:
        for offset, item in enumerate(reservations, start=0):
            row = data_start + offset
            if offset > 0:
                ws.insert_rows(row)
            _copy_row_style(ws, style_row, row, max_col)
            for col, value in enumerate(reservation_to_open_record(offset + 1, item), start=1):
                ws.cell(row=row, column=col).value = value
    else:
        ws.insert_rows(data_start)
        _copy_row_style(ws, style_row, data_start, max_col)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
